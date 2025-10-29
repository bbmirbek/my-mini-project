import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

def format_and_save_report(result_df, fines_df, summary_df, pre_last_df, last_df, 
                           correction,  path):
    # 1) Выравниваем длины и объединяем
    max_len = max(len(result_df), len(fines_df), len(summary_df), len(pre_last_df), len(last_df), 1)
    result_df  = result_df.reindex(range(max_len))
    fines_df   = fines_df.reindex(range(max_len))
    summary_df = summary_df.reindex(range(max_len))  # первая строка с цифрами, ниже пусто
    pre_last_df = pre_last_df.reindex(range(max_len))
    last_df = last_df.reindex(range(max_len))
                                    

    combined = pd.concat([result_df, fines_df, summary_df, pre_last_df, last_df], axis=1)

    corrections = {"Артикул поставщика": "Корректировка"}
    corrections["Выручка (продажи - возвраты)"] = correction[0]
    corrections["Комиссия WB"] = correction[0]
    corrections["Сумма к перечислению"] = correction[1]
    corrections["Реклама с собственного счёта"] = correction[2]

    corrections_row = pd.DataFrame(corrections)
    combined = pd.concat([combined, corrections_row], ignore_index=True)

    # 2) Гарантируем числовой тип где возможно (важно для сумм)
    for col in combined.columns:
        if col not in ("Артикул поставщика", "Виды штрафов", "Себестоимость единицы товара"):
            combined[col] = pd.to_numeric(combined[col], errors="coerce")


    # 3) TOTAL (учтёт и summary_df, т.к. всё уже вместе)
    totals = combined.select_dtypes(include="number").sum(numeric_only=True)
    totals["Реклама с собственного счёта"] -= correction[2]
    totals_row = pd.DataFrame([{**{"Артикул поставщика": "Total:"}, **totals.to_dict()}])
    combined = pd.concat([combined, totals_row], ignore_index=True)

    # 4) PERCENTAGE (от общей выручки)
    revenue_col = "Выручка (продажи - возвраты)"
    total_rev = float(totals.get(revenue_col, 0) or 0)

    exclude_pct = {
        "Артикул поставщика",
        "Виды штрафов",
        "Кол-во продаж",
        "Себестоимость единицы товара",
    }

    pct_dict = {"Артикул поставщика": "Percentage:"}

    if total_rev != 0:
        for col in combined.columns:
            if col in exclude_pct:
                if col == "Артикул поставщика":
                    continue
                pct_dict[col] = ""
                continue
            val = totals.get(col, None)
            pct_dict[col] = f"{round(val / total_rev * 100, 2)}%" if pd.notna(val) else ""
    else:
        for col in combined.columns:
            pct_dict[col] = "" if col != "Артикул поставщика" else "Percentage:"

    percentages_row = pd.DataFrame([pct_dict])
    combined = pd.concat([combined, percentages_row], ignore_index=True)

    # 5) Заменяем NaN только перед записью (чтобы не ломать типы)
    combined = combined.fillna("")

    # 6) Сохранение
    combined.to_excel(path, index=False, sheet_name="Отчёт")

    # 7) Лёгкое форматирование
    wb = load_workbook(path)
    ws = wb.active

    header_fill = PatternFill("solid", fgColor="C0C0C0")
    header_font = Font(bold=True, color="FF000000")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="FF000000")
    border = Border(left=thin, right=thin,top=thin,bottom=thin)

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bej_fill = PatternFill(start_color="F3DFD7", end_color="F3DFD7", fill_type="solid")
    blue_fill = PatternFill(start_color="3EA6E8", end_color="3EA6E8", fill_type="solid")
    green_fill = PatternFill(start_color="77DD77", end_color="77DD77", fill_type="solid")
    red_fill = PatternFill(start_color="F9A6A6", end_color="F9A6A6", fill_type="solid")

    # закрашиваем диапазон B2:D4
    for row in ws[f"P2:P{ws.max_row - 3}"]:
        for cell in row:
            cell.fill = yellow_fill
            cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 1, max_col=ws.max_column):
        for cell in row:
            col_letter = cell.column_letter
            if col_letter in ["C", "D", "E", "F", "G", "H", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S"]:
                cell.number_format = numbers.FORMAT_NUMBER_00

    ws.merge_cells(f"J2:J{ws.max_row-3}")
    ws.merge_cells(f"K2:K{ws.max_row-3}")
    ws.merge_cells(f"L2:L{ws.max_row-3}")
    ws.merge_cells(f"M2:M{ws.max_row-3}")
    ws.merge_cells(f"N2:N{ws.max_row-3}")
    ws.merge_cells(f"O2:O{ws.max_row-3}")
    ws.merge_cells(f"S2:S{ws.max_row-3}")


    ws.insert_rows(idx=ws.max_row - 2, amount=3)
    ws.insert_rows(idx=ws.max_row - 1, amount=1)
    ws.insert_rows(idx=ws.max_row, amount=1)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = center
            cell.border = border

    # автоширина
    ws.row_dimensions[1].height = 35
    for column in ws.columns:
        col_letter = column[0].column_letter
        max_len = 0
        for cell in column:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(max_len, 10) + 2

        if col_letter in ["A", "B"]:
            ws[f"{col_letter}{ws.max_row}"].fill = bej_fill
            ws[f"{col_letter}{ws.max_row - 2}"].fill = bej_fill
        elif col_letter == "S":
            ws[f"{col_letter}{ws.max_row}"].fill = blue_fill
            ws[f"{col_letter}{ws.max_row - 2}"].fill = blue_fill
        elif col_letter in ["C", "F", "N"]:
            ws[f"{col_letter}{ws.max_row}"].fill = green_fill
            ws[f"{col_letter}{ws.max_row - 2}"].fill = green_fill
        else:
            ws[f"{col_letter}{ws.max_row}"].fill = red_fill
            ws[f"{col_letter}{ws.max_row - 2}"].fill = red_fill

    ws[f"O{ws.max_row - 4}"].fill = yellow_fill

    ws.freeze_panes = "D1"

    wb.save(path)
