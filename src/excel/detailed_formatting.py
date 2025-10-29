import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

def format_and_save_detailed_report(result_df, correction, Buyout, path):
    corrections = {"Артикул поставщика" : "Корректировка"}
    corrections["Выручка (продажи - возвраты)"] = correction[0]
    corrections["Комиссия WB"] = correction[1]
    corrections["Сумма к перечислению"] = correction[2]

    corrections_row = pd.DataFrame(corrections)
    result_df = pd.concat([result_df, corrections_row], ignore_index=True)

    not_numeric = ["Артикул поставщика", "Выручка %", "Логистика %", "Хранение % от собственного дохода",
                   "Хранение % от всей суммы", "Реклама %", "Себестоимость единицы товара",
                   "Себестоимость %", "Чистая Прибыль", "Чистая Прибыль %",  "Выкуп"]

    # 2) Гарантируем числовой тип где возможно (важно для сумм)    
    for col in result_df.columns:
        if col not in not_numeric:
            result_df[col] = pd.to_numeric(result_df[col], errors="coerce")

    # 3) TOTAL (учтёт и summary_df, т.к. всё уже вместе)
    totals = result_df.select_dtypes(include="number").sum(numeric_only=True)
    totals["Чистая Прибыль"] = (totals["Сумма к перечислению"] - totals["Реклама"] - totals["Общая себестоимость"] -
                                totals["Логистика"] - totals["Upsell-услуги (5%)"] - totals["Штрафы"] - 
                                totals["Хранение на складе"] - totals["Джем"])
    totals["Выкуп"] = f"{round(Buyout, 2)}%"
    totals_row = pd.DataFrame([{**{"Артикул поставщика": "Total:"}, **totals.to_dict()}])
    result_df = pd.concat([result_df, totals_row], ignore_index=True)

    # 4) PERCENTAGE (от общей выручки)
    revenue_col = "Выручка (продажи - возвраты)"
    total_rev = float(totals.get(revenue_col, 0) or 0)

    exclude_pct = {
        "Артикул поставщика",
        "Кол-во продаж",
        "Выручка %",
        "Логистика %",
        "Хранение % от собственного дохода",
        "Хранение % от всей суммы",
        "Реклама %",
        "Себестоимость единицы товара",
        "Себестоимость %",
        "Чистая Прибыль %",
        "Выкуп"
    }

    pct_dict = {"Артикул поставщика": "Percentage:"}

    if total_rev != 0:
        for col in result_df.columns:
            if col in exclude_pct:
                if col == "Артикул поставщика":
                    continue
                pct_dict[col] = ""
                continue
            val = totals.get(col, None)
            pct_dict[col] = f"{round(val / total_rev * 100, 2)}%" if pd.notna(val) else ""
    else:
        for col in result_df.columns:
            pct_dict[col] = "" if col != "Артикул поставщика" else "Percentage:"

    percentages_row = pd.DataFrame([pct_dict])
    result_df = pd.concat([result_df, percentages_row], ignore_index=True)

    # 5) Заменяем NaN только перед записью (чтобы не ломать типы)
    result_df = result_df.fillna("")

    # 6) Сохранение
    result_df.to_excel(path, index=False, sheet_name="Отчёт")

    # 7) Лёгкое форматирование
    wb = load_workbook(path)
    ws = wb.active

    header_fill = PatternFill("solid", fgColor="C0C0C0")
    header_font = Font(bold=True, color="FF000000")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="FF000000")
    border = Border(left=thin, right=thin,top=thin,bottom=thin)

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bej_fill = PatternFill(start_color="F3DFD7", end_color="F3DFD7", fill_type="solid")
    blue_fill = PatternFill(start_color="3EA6E8", end_color="3EA6E8", fill_type="solid")
    green_fill = PatternFill(start_color="77DD77", end_color="77DD77", fill_type="solid")
    red_fill = PatternFill(start_color="F9A6A6", end_color="F9A6A6", fill_type="solid")
    bblue_fill = PatternFill(start_color="76B9CE", end_color="76B9CE", fill_type="solid")

    for row in ws[f"S2:S{ws.max_row - 3}"]:
        for cell in row:
            cell.fill = yellow_fill
            cell.border = border
    for row in ws[f"Q2:Q{ws.max_row - 3}"]:
        for cell in row:
            cell.fill = yellow_fill
            cell.border = border

    ws.insert_rows(idx=ws.max_row - 2, amount=3)
    ws.insert_rows(idx=ws.max_row - 1, amount=1)
    ws.insert_rows(idx=ws.max_row, amount=1)

    for cell in ws[1]:
        col_letter = cell.column_letter

        if col_letter not in ["D", "I", "K", "L", "X", "U", "R"]:
            cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = center
            cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 1, max_col=ws.max_column):
        for cell in row:
            col_letter = cell.column_letter
            if col_letter in ["C", "E", "F", "G", "H", "J", "M", "N", "O", "P", "Q", "T", "V", "W"]:
                cell.number_format = numbers.FORMAT_NUMBER_00
            
    ws.row_dimensions[1].height = 35
    for column in ws.columns:
        col_letter = column[0].column_letter
        max_len = 0
        for cell in column:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(max_len, 10) + 2

        if col_letter in ["A", "B"]:
            ws[f"{col_letter}{ws.max_row}"].fill = bej_fill
            ws[f"{col_letter}{ws.max_row - 2}"].fill = bej_fill
        elif col_letter == "W":
            ws[f"{col_letter}{ws.max_row}"].fill = blue_fill
            ws[f"{col_letter}{ws.max_row - 2}"].fill = blue_fill
        elif col_letter in ["C", "G", "P"]:
            ws[f"{col_letter}{ws.max_row}"].fill = green_fill
            ws[f"{col_letter}{ws.max_row - 2}"].fill = green_fill
        elif col_letter in ["E", "F", "H", "J", "M", "N", "O", "Q", "T", "V"]:
            ws[f"{col_letter}{ws.max_row}"].fill = red_fill
            ws[f"{col_letter}{ws.max_row - 2}"].fill = red_fill
        elif col_letter in ["Y"]:
            ws[f"{col_letter}{ws.max_row - 2}"].fill = bblue_fill

    ws.freeze_panes = "D1"
    
    wb.save(path)

    print("✅ detailed ")
